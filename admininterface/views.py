from django.shortcuts import render,redirect
from .models import Product,Cart,CartItem,Table,Courier
from django.views.generic.edit import CreateView,UpdateView
from django.views.generic.list import ListView
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect,HttpResponse, Http404
import os
from django.conf import settings
from django.http import JsonResponse
import json
from datetime import date,timedelta
from django.core import serializers
import openpyxl
from openpyxl.styles.borders import Border, Side
import os




def alert(request,**kwargs):
    amount=Cart.objects.all().count()
    return JsonResponse(data={
        'amount': amount
    })


def main(request):
    if request.POST:
        courier=Courier.objects.create(name=request.POST.get('courier'))
    if request.GET.get('order'):
        pay_method=request.GET.get('pay_method')
        cart=Cart.objects.create(customer='Ayakta Müşteri',address='None',number='(0212) 345 67 89',pay_method=pay_method,status=4)
        for i in json.loads(request.GET.get('order')):
            product=Product.objects.get(name=i['product'])
            quantity=i['quantity']
            total_price=i['price']*quantity
            CartItem.objects.create(product=product,quantity=quantity,total_price=total_price,cart=cart)
    products=Product.objects.all()
    return render(request,'admininterface/home.html',{'products':products})


def chart(request):
    products=[]
    amount=[]
    today = date.today()
    for i in Product.objects.all().values_list('name'):
        product=Product.objects.filter(name=str(i)[2:-3])
        if CartItem.objects.filter(product__in=product,
                                           created_on__year=today.year,
                                           created_on__month=today.month,
                                           created_on__day=today.day).count()==0:
            continue
        sum=0
        for j in CartItem.objects.filter(product__in=product,
                                             created_on__year=today.year,
                                             created_on__month=today.month,
                                             created_on__day=today.day):
            sum+=j.quantity
        products.append(str(i)[2:-3])
        amount.append(sum)
    return JsonResponse(data={
        'labels': products,
        'data': amount,
    })


def download(request):
        file_path = os.path.join(settings.BASE_DIR,"indir.xlsx")
        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                return response
        raise Http404


def excel(request):
    try:
        os.remove("indir.xlsx")
    except:
        pass
    wb=openpyxl.Workbook()
    ws=wb.active
    headers=['Ürün','Adet','Müşteri','Tarih','Fiyat','Kurye']
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    for x in range(6):
        ws.cell(row=1,column=x+1).value=headers[x]
        ws.cell(row=1,column=x+1).border=thin_border
    k=0
    sum=0
    for i in request.GET.lists():
        k+=1
        l=1
        for j in i[1]:
            l+=1
            if k==5:
                ws.cell(row=l,column=k).value=int(j)
                sum+=int(j)
            else:
                ws.cell(row=l,column=k).value=j
            ws.cell(row=l,column=k).border=thin_border
    ws.cell(row=l+2,column=4).value='TOPLAM:'
    ws.cell(row=l+2,column=5).value=sum
    wb.save("indir.xlsx")
    return HttpResponse()


def sell(request):
    table=Table.objects.all()
    if request.GET.get('name'):
        cart=Cart.objects.create(customer=request.GET.get('name'),address='None',number='(0212) 345 67 89',status=4)
        table=Table.objects.create(name=request.GET.get('name'),cart=cart)
    tables=Table.objects.all()
    return render(request,'admininterface/sell.html',{'tables':tables})

def history(request):
    today = date.today()
    list=CartItem.objects.filter(created_on__year=today.year,
                                       created_on__month=today.month,
                                       created_on__day=today.day).order_by('-created_on')
    couriers1=Courier.objects.all()
    products1=Product.objects.all()
    customers1=Cart.objects.all().values_list('customer',flat=True).distinct()
    if request.GET.get('courier'):
        customer=request.GET.get('customer')
        product=request.GET.get('product')
        courier=request.GET.get('courier')
        date1=request.GET.get('date1')
        date2=request.GET.get('date2')
        carts=[]
        products=[]
        couriers=[]
        cartitems=[]
        if date1!='':
            d2=date(int(date2[:4]),int(date2[5:7]),int(date2[8:]))
            d1=date(int(date1[:4]),int(date1[5:7]),int(date1[8:]))
            dd=CartItem.objects.filter(created_on__range=[d1, d2]).order_by('-created_on')
        else:
            dd=CartItem.objects.all().order_by('-created_on')
        if customer!='None':
            ci=CartItem.objects.filter(cart__customer__contains=customer).order_by('-created_on')
        else:
            ci=dd
        if product!='None':
            p=CartItem.objects.filter(product__name__contains=product).order_by('-created_on')
        else:
            p=dd
        if courier!='None':
            c=CartItem.objects.filter(cart__courier__name__contains=courier).order_by('-created_on')
        else:
            c=dd
        try:
            intersection=ci&p&c&dd
            for i in intersection:
                cartitems.append(i)
                carts.append(i.cart)
                products.append(i.product)
                couriers.append(i.cart.courier)
            return JsonResponse(data={
                'cartitems': serializers.serialize("json", cartitems),
                'carts': serializers.serialize("json",carts),
                'products': serializers.serialize("json",products),
                'couriers': serializers.serialize("json",couriers),
            })
        except:
            pass
    return render(request,'admininterface/history.html',{'list':list,'couriers':couriers1,'products':products1,'customers':customers1})




def table_view(request,pk):
    table=Table.objects.get(id=pk)
    cart=Cart.objects.get(id=table.cart.id)
    products=CartItem.objects.filter(cart=cart)
    realproducts=Product.objects.all()
    if request.GET.get('change'):
        table=Table.objects.get(id=request.GET.get('table_id'))
        cart=Cart.objects.get(id=table.cart.id)
        j=0
        for i in request.GET.lists():
            if ('table_id' not in i) and ('change' not in i):
                if j%2==0:
                    product=Product.objects.get(name=i[1][0])
                else:
                    quantity=int(i[1][0])
                    CartItem.objects.update_or_create(
                    product=product,cart=cart,defaults={'total_price':quantity*product.price,'quantity':quantity})
                j+=1
    if request.GET.get('payment'):
        table=Table.objects.get(id=request.GET.get('table_id'))
        cart=Cart.objects.get(id=table.cart.id)
        cart.pay_method=request.GET.get('payment')
        cart.status=5
        cart.save()
        cart_new=Cart.objects.create(customer=table.name,address='None',number='(0212) 345 67 89',status=4)
        table.cart=cart_new
        table.save()
    return render(request,'admininterface/table.html',{'table':table,'products':products,'realproducts':realproducts})


class CreateProduct(CreateView):
    template_name="admininterface/urunler_form.html"
    model=Product
    fields='__all__'
    success_url='/yonetici/urun'

    def get_context_data(self,**kwargs):
        context=super().get_context_data(**kwargs)
        context['products']=Product.objects.order_by('-id')
        return context



class UpdateProduct(UpdateView):
    model=Product
    fields='__all__'
    template_name='admininterface/product_update_form.html'
    success_url='/yonetici/urun'

    def get_form(self, form_class=None):
        form = super(UpdateProduct, self).get_form(form_class)
        form.fields['name'].widget.attrs.update({'class':'form-control'})
        form.fields['price'].widget.attrs.update({'class': 'form-control'})
        form.fields['image'].widget.attrs.update({'class': 'form-control'})
        form.fields['name'].label='Ürün Adı'
        form.fields['image'].label='Resim'
        form.fields['price'].label='Fiyat'
        form.fields['status'].label='Durum'
        return form



def delete(request,pk):
    object=Product.objects.get(pk=pk)
    object.delete()
    return HttpResponseRedirect('/yonetici/urun')



class ViewCart(ListView):
    model=Cart
    field='__all__'
    template_name='admininterface/view_cart.html'
    success_url='/yonetici/siparis'

    def get(self,request,**kwargs):
        if request.GET.get('courier'):
            courier=Courier.objects.get(name=request.GET.get('courier'))
            cart=Cart.objects.get(id=int(request.GET.get('id')))
            cart.courier=courier
            cart.save()
        if request.GET.get('status'):
            cart_change=Cart.objects.get(id=int(request.GET.get('id')))
            cart_change.status=int(request.GET.get('status'))+1
            cart_change.save()
            return redirect('admininterface:view_cart')
        j=0
        for i in request.GET.lists():
            if 'delete' in i[0]:
                cartitem_delete=CartItem.objects.get(id=i[1][0]).delete()
            else:
                if j%4==0:
                    product=Product.objects.get(name=i[1][0])
                elif j%4==1:
                    total_price=i[1][0]
                elif j%4==2:
                    quantity=i[1][0]
                else:
                    cart=Cart.objects.get(id=i[1][0])
                    cartitem,created=CartItem.objects.update_or_create(
                    product=product,cart=cart,defaults={'total_price':total_price,'quantity':quantity})
                j+=1
        carts=Cart.objects.order_by('-created_on')[:100]
        products=Product.objects.order_by('-name')
        couriers=Courier.objects.all()
        return render(request,'admininterface/view_cart.html',{'carts':carts,'products':products,'couriers':couriers})
